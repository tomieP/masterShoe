from django.core.management.base import BaseCommand, CommandError
from django.db import transaction, IntegrityError
import os
import sys
from openpyxl import load_workbook
from management.models import Product, ProductVariant


class Command(BaseCommand):
    help = 'Import products and variants from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file (.xlsx)')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse file without making changes to database',
        )

    def safe_stdout(self, message, style_func=None):
        """Safely write message to stdout handling encoding issues"""
        try:
            # Try to encode with console encoding, replace errors with '?'
            encoding = sys.stdout.encoding or 'utf-8'
            safe_msg = message.encode(encoding, errors='replace').decode(encoding)

            if style_func:
                self.stdout.write(style_func(safe_msg))
            else:
                self.stdout.write(safe_msg)
        except Exception:
            # Fallback for very restrictive environments
            safe_msg = ''.join(c if ord(c) < 128 else '?' for c in message)
            if style_func:
                self.stdout.write(style_func(safe_msg))
            else:
                self.stdout.write(safe_msg)

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']

        # Validate file exists
        if not os.path.exists(file_path):
            raise CommandError(f'File "{file_path}" does not exist')

        # Validate file extension
        if not file_path.lower().endswith('.xlsx'):
            raise CommandError('File must be an Excel file (.xlsx)')

        self.safe_stdout(
            f'Processing {"dry run of " if dry_run else ""}"{file_path}"...',
            self.style.SUCCESS
        )

        try:
            # Load workbook
            wb = load_workbook(filename=file_path, read_only=True)
            ws = wb.active

            # Get header row
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Map Vietnamese column names to expected fields
            column_mapping = {
                'Tên': 'name',
                'Thương hiệu': 'brand',
                'Size': 'size',
                'Màu': 'color',
                'Giá nhập': 'import_price',
                'Giá bán': 'selling_price',
                'Tồn': 'quantity'
            }

            # Find column indices
            col_indices = {}
            missing_columns = []

            for vi_name, eng_name in column_mapping.items():
                try:
                    col_idx = header.index(vi_name)
                    col_indices[eng_name] = col_idx
                except ValueError:
                    missing_columns.append(vi_name)

            if missing_columns:
                raise CommandError(
                    f'Missing required columns: {", ".join(missing_columns)}. '
                    f'Expected columns: {", ".join(column_mapping.keys())}'
                )

            # Statistics
            stats = {
                'products_created': 0,
                'products_found': 0,
                'variants_created': 0,
                'variants_updated': 0,
                'rows_processed': 0,
                'errors': []
            }

            # Process rows
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                stats['rows_processed'] += 1

                try:
                    # Extract cell values
                    row_data = {}
                    for field_name, col_idx in col_indices.items():
                        cell_value = row[col_idx].value
                        row_data[field_name] = cell_value

                    # Validate required fields
                    if not all([row_data['name'], row_data['brand'],
                               row_data['size'], row_data['color']]):
                        stats['errors'].append(
                            f'Row {row_num}: Missing required fields (name, brand, size, or color)'
                        )
                        continue

                    # Validate numeric fields
                    try:
                        import_price = float(row_data['import_price']) if row_data['import_price'] is not None else 0
                        selling_price = float(row_data['selling_price']) if row_data['selling_price'] is not None else 0
                        quantity = int(row_data['quantity']) if row_data['quantity'] is not None else 0

                        if import_price < 0 or selling_price < 0:
                            raise ValueError("Prices cannot be negative")

                    except (ValueError, TypeError) as e:
                        stats['errors'].append(
                            f'Row {row_num}: Invalid price/quantity format - {str(e)}'
                        )
                        continue

                    # Use transaction for each row to ensure consistency
                    with transaction.atomic():
                        # Find or create Product
                        product_name = str(row_data['name']).strip()
                        product_brand = str(row_data['brand']).strip()

                        product, product_created = Product.objects.get_or_create(
                            name=product_name,
                            brand=product_brand,
                            defaults={
                                'code': self.generate_product_code(product_name, product_brand),
                                'type': 'Chung',  # Default type
                                'is_active': True
                            }
                        )

                        if product_created:
                            stats['products_created'] += 1
                            if not dry_run:
                                self.safe_stdout(
                                    f'[+] Row {row_num}: Created Product "{product_name}" (Brand: {product_brand})',
                                    self.style.SUCCESS
                                )
                        else:
                            stats['products_found'] += 1

                        # Generate SKU
                        sku = self.generate_sku(product.code, str(row_data['color']), str(row_data['size']))

                        # Find or create ProductVariant
                        variant_color = str(row_data['color']).strip()
                        variant_size = str(row_data['size']).strip()

                        variant_defaults = {
                            'import_price': import_price,
                            'selling_price': selling_price,
                            'sku': sku,
                            'min_quantity': 5,  # Default minimum quantity
                            'is_active': True
                        }

                        if not dry_run:
                            variant, variant_created = ProductVariant.objects.update_or_create(
                                product=product,
                                color=variant_color,
                                size=variant_size,
                                defaults=variant_defaults
                            )

                            # Update stock quantity (add to existing or set new)
                            if variant_created:
                                variant.stock_quantity = quantity
                                variant.save()
                                stats['variants_created'] += 1
                                action = "Created"
                            else:
                                variant.stock_quantity += quantity  # Increment existing stock
                                variant.save()
                                stats['variants_updated'] += 1
                                action = "Updated stock"
                        else:
                            # Dry run - just check if would create/update
                            variant_exists = ProductVariant.objects.filter(
                                product=product,
                                color=variant_color,
                                size=variant_size
                            ).exists()

                            if variant_exists:
                                stats['variants_updated'] += 1
                                action = "Would update stock"
                            else:
                                stats['variants_created'] += 1
                                action = "Would create"

                        if not dry_run:
                            self.safe_stdout(
                                f'[+] Row {row_num}: {action} Variant (Size: {variant_size}, Color: {variant_color})',
                                self.style.SUCCESS
                            )

                except Exception as e:
                    stats['errors'].append(f'Row {row_num}: Unexpected error - {str(e)}')
                    continue

            wb.close()

            # Print summary
            self.stdout.write('')  # Empty line
            if dry_run:
                self.safe_stdout(
                    f'[DRY RUN COMPLETED] Would create {stats["products_created"]} products, '
                    f'{stats["variants_created"]} variants and update {stats["variants_updated"]} variants',
                    self.style.WARNING
                )
            else:
                self.safe_stdout(
                    f'[OK] Successfully imported {stats["products_created"]} products and '
                    f'{stats["variants_created"]} variants (updated {stats["variants_updated"]} existing variants)',
                    self.style.SUCCESS
                )

            # Print errors if any
            if stats['errors']:
                self.stdout.write('')  # Empty line
                self.safe_stdout(f'Encountered {len(stats["errors"])} errors:', self.style.WARNING)
                for error in stats['errors'][:10]:  # Show first 10 errors
                    self.safe_stdout(f'  - {error}')
                if len(stats['errors']) > 10:
                    self.safe_stdout(f'  ... and {len(stats["errors"]) - 10} more errors')

        except Exception as e:
            raise CommandError(f'Error processing Excel file: {str(e)}')

    def generate_product_code(self, name, brand):
        """Generate a unique product code based on name and brand"""
        import hashlib
        from datetime import datetime

        # Create a hash from name and brand
        hash_input = f"{name}_{brand}_{datetime.now().strftime('%Y%m%d')}"
        hash_object = hashlib.md5(hash_input.encode('utf-8'))
        hash_hex = hash_object.hexdigest()[:8].upper()

        # Take first 3 letters of name and brand, strip non-alpha
        name_part = ''.join([c for c in name[:5] if c.isalnum()]).upper() or 'PROD'
        brand_part = ''.join([c for c in brand[:3] if c.isalnum()]).upper() or 'BRD'

        return f"{name_part}{brand_part}{hash_hex}"

    def generate_sku(self, product_code, color, size):
        """Generate SKU for product variant"""
        # Clean color and size for SKU
        color_clean = ''.join(c for c in color if c.isalnum())[:3].upper() or 'COL'
        size_clean = ''.join(c for c in size if c.isalnum())[:3] or 'SZ'

        return f"{product_code}-{color_clean}-{size_clean}"